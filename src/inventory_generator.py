import csv
import io
import json
import os.path
import statistics
from typing import List

import pandas as pd
import openpyxl
from omegaconf import DictConfig
from openai import OpenAI
from openai.types.beta import Thread, Assistant

from src.models.message import Message
from src.models.consolidated_sheet import ConsolidatedSheetItem
from src.models.report_sheet_item import ReportSheetItem
from src.models.separate_sheet_metadata import SeparateSheetMetadata
from src.price_generator import get_price_results
from src.report_generator import generate_report

client: OpenAI = None


def generate_inventory(
    output_directory: str,
    messages: List[Message],
    system_logger: "loguru.system_logger",
    config: DictConfig,
):
    global client
    client = OpenAI(api_key=config.secrets.openai_api_key)
    assistant = initiate_assistant()
    thread = None

    report_sheet_items: List[ReportSheetItem] = []
    separate_sheets: List[SeparateSheetMetadata] = []
    consolidated_sheet_items = []

    message_counter = 0
    total_messages = len(messages)
    for message in messages:
        message_counter += 1
        print(f"⏳ Analysing {message_counter}/{total_messages} messages")

        attachment_counter = 0
        for attachment in message.attachments:

            # Currently supported attachments: .xls or .xlsx
            if attachment.name.endswith(".xlsx") or attachment.name.endswith(".xls"):
                attachment_counter += 1
                sheets = {}
                attachment_path = os.path.join(
                    output_directory,
                    "Messages",
                    f"Message {message_counter}",
                    "Attachments",
                    attachment.name,
                )
                try:
                    try:
                        workbook = openpyxl.load_workbook(
                            attachment_path, data_only=True
                        )
                    except Exception as ex:
                        if str(ex).startswith(
                            "openpyxl does not support the old .xls file format"
                        ):
                            xls = pd.ExcelFile(attachment_path, engine="xlrd")
                            with pd.ExcelWriter(
                                os.path.join("temp", "temp.xlsx"), engine="openpyxl"
                            ) as writer:
                                for xls_sheet_name in xls.sheet_names:
                                    df = pd.read_excel(
                                        xls, sheet_name=xls_sheet_name, engine="xlrd"
                                    )
                                    df.to_excel(
                                        writer, sheet_name=xls_sheet_name, index=False
                                    )
                            attachment_path = os.path.join("temp", "temp.xlsx")
                            workbook = openpyxl.load_workbook(
                                attachment_path, data_only=True
                            )
                        else:
                            raise

                    # Try removing hidden rows from all the sheets
                    try:
                        for sheet_name in workbook.sheetnames:
                            sheet = workbook[sheet_name]
                            hidden_rows = []
                            for row in sheet.row_dimensions:
                                if sheet.row_dimensions[row].hidden:
                                    hidden_rows.append(row - 1)
                            df = pd.read_excel(
                                attachment_path,
                                sheet_name=sheet.title,
                                engine="openpyxl",
                                header=None,
                            )
                            sheets[sheet_name] = df.drop(hidden_rows, axis=0)
                    except:
                        pass

                    if os.path.exists(os.path.join("temp", "temp.xlsx")):
                        os.remove(os.path.join("temp", "temp.xlsx"))

                    sheet_counter = 0
                    for sheet_name, df in sheets.items():
                        csv_output = io.StringIO()
                        df.to_csv(csv_output, index=False)
                        csv_string = csv_output.getvalue()
                        if csv_string.strip():
                            csv_string_io = io.StringIO(csv_string)
                            csv_reader = csv.reader(
                                csv_string_io, delimiter=",", quotechar='"'
                            )

                            sheet_counter += 1
                            separate_sheet_metadata: SeparateSheetMetadata = (
                                SeparateSheetMetadata()
                            )
                            separate_sheet_metadata.name = f"M-{message_counter}A-{attachment_counter}S-{sheet_counter}"
                            report_sheet_item: ReportSheetItem = ReportSheetItem(
                                sender=message.sender.name
                                + " - "
                                + message.sender.address,
                                received_at=message.received_at,
                                sheet_name=separate_sheet_metadata.name,
                                status="",
                                comments="",
                                file_name=attachment.name,
                                file_path=f"{os.path.join(output_directory, 'Messages', 'Message ' + str(message_counter), 'Attachments', attachment.name)}",
                            )
                            separate_sheet = []
                            table_matrix_indices = {"left": -1, "top": -1, "right": -1}
                            header_column_index = {
                                "barcode": -1,
                                "quantity": -1,
                                "product": -1,
                                "price": -1,
                            }

                            # GPT header analysis
                            # First 30 rows
                            thread, analysis = analyse(
                                df.head(30).to_csv(index=False), thread, assistant
                            )
                            if analysis.status == "completed":
                                response_messages = client.beta.threads.messages.list(
                                    thread_id=thread.id
                                )
                                # Sample
                                # response = {"barcode": "EAN", "quantity": "", "product": "PRODUCTS", "price": "PRICE €"}
                                try:
                                    response = json.loads(
                                        str(
                                            response_messages.data[0]
                                            .content[0]
                                            .text.value
                                        )
                                        .strip()
                                        .strip("```json")
                                        .strip()
                                    )
                                except:
                                    response = None

                                # No required headers detected
                                if response is None or (
                                    response["barcode"].strip() == ""
                                    and response["quantity"].strip() == ""
                                    and response["product"].strip() == ""
                                    and response["price"].strip() == ""
                                ):
                                    # First 50 rows
                                    thread, analysis = analyse(
                                        df.head(50).to_csv(index=False),
                                        thread,
                                        assistant,
                                    )
                                    if analysis.status == "completed":
                                        response_messages = (
                                            client.beta.threads.messages.list(
                                                thread_id=thread.id
                                            )
                                        )
                                        response = json.loads(
                                            str(
                                                response_messages.data[0]
                                                .content[0]
                                                .text.value
                                            )
                                            .strip()
                                            .strip("```json")
                                            .strip()
                                        )

                                # At least one required header detected
                                if (
                                    response["barcode"].strip() != ""
                                    or response["quantity"].strip() != ""
                                    or response["product"].strip() != ""
                                    or response["price"].strip() != ""
                                ):
                                    header_row_found = False

                                    for row_index, row in enumerate(csv_reader):
                                        columns = [
                                            str(col).strip()
                                            for col_idx, col in enumerate(row)
                                        ]

                                        if not header_row_found:

                                            # If any column matches with the required header value, reset the results obtained from previous rows
                                            if (
                                                (
                                                    response["barcode"].strip() != ""
                                                    and response["barcode"].strip()
                                                    in columns
                                                )
                                                or (
                                                    response["quantity"].strip() != ""
                                                    and response["quantity"].strip()
                                                    in columns
                                                )
                                                or (
                                                    response["product"].strip() != ""
                                                    and response["product"].strip()
                                                    in columns
                                                )
                                                or (
                                                    response["price"].strip() != ""
                                                    and response["price"].strip()
                                                    in columns
                                                )
                                            ):
                                                header_column_index = {
                                                    "barcode": -1,
                                                    "quantity": -1,
                                                    "product": -1,
                                                    "price": -1,
                                                }

                                            if (
                                                response["barcode"].strip() != ""
                                                and response["barcode"].strip()
                                                in columns
                                            ):
                                                header_column_index["barcode"] = (
                                                    columns.index(
                                                        response["barcode"].strip()
                                                    )
                                                )
                                            if (
                                                response["quantity"].strip() != ""
                                                and response["quantity"].strip()
                                                in columns
                                            ):
                                                header_column_index["quantity"] = (
                                                    columns.index(
                                                        response["quantity"].strip()
                                                    )
                                                )
                                            if (
                                                response["product"].strip() != ""
                                                and response["product"].strip()
                                                in columns
                                            ):
                                                header_column_index["product"] = (
                                                    columns.index(
                                                        response["product"].strip()
                                                    )
                                                )
                                            if (
                                                response["price"].strip() != ""
                                                and response["price"].strip() in columns
                                            ):
                                                header_column_index["price"] = (
                                                    columns.index(
                                                        response["price"].strip()
                                                    )
                                                )

                                            # Required number of headers to detect a header row: 4, OK to also have: >50% (2 or more)
                                            existence_count = len(
                                                [
                                                    index
                                                    for index in header_column_index.values()
                                                    if index != -1
                                                ]
                                            )
                                            if existence_count >= 2:
                                                header_row_found = True

                                                # For separate sheet
                                                table_matrix_indices["top"] = row_index
                                                calculate_table_matrix_indices(
                                                    table_matrix_indices,
                                                    [
                                                        row_enum
                                                        for row_enum_index, row_enum in enumerate(
                                                            csv.reader(
                                                                io.StringIO(csv_string),
                                                                delimiter=",",
                                                                quotechar='"',
                                                            )
                                                        )
                                                    ],
                                                    header_column_index,
                                                )
                                                result_columns = []
                                                for column in columns[
                                                    table_matrix_indices[
                                                        "left"
                                                    ] : table_matrix_indices["right"]
                                                    + 1
                                                ]:
                                                    if (
                                                        str(column)
                                                        .strip()
                                                        .startswith("Unnamed:")
                                                    ):
                                                        result_columns.append("")
                                                    else:
                                                        result_columns.append(column)
                                                separate_sheet.append(result_columns)
                                        else:
                                            # For consolidated sheet
                                            consolidated_sheet_item: (
                                                ConsolidatedSheetItem
                                            ) = ConsolidatedSheetItem(
                                                sender=message.sender.name
                                                + " - "
                                                + message.sender.address,
                                                barcode="",
                                                quantity="",
                                                product_description="",
                                                unit_price="",
                                            )
                                            if header_column_index["barcode"] != -1:
                                                try:
                                                    consolidated_sheet_item.barcode = (
                                                        str(
                                                            columns[
                                                                header_column_index[
                                                                    "barcode"
                                                                ]
                                                            ]
                                                        ).strip()
                                                    )
                                                    try:
                                                        consolidated_sheet_item.barcode = int(
                                                            float(
                                                                consolidated_sheet_item.barcode
                                                            )
                                                        )
                                                    except:
                                                        pass
                                                except:
                                                    # May indicate a noisy row
                                                    pass
                                            if header_column_index["quantity"] != -1:
                                                try:
                                                    consolidated_sheet_item.quantity = (
                                                        str(
                                                            columns[
                                                                header_column_index[
                                                                    "quantity"
                                                                ]
                                                            ]
                                                        ).strip()
                                                    )
                                                    try:
                                                        consolidated_sheet_item.quantity = int(
                                                            float(
                                                                consolidated_sheet_item.quantity
                                                            )
                                                        )
                                                    except:
                                                        pass
                                                except:
                                                    # May indicate a noisy row
                                                    pass
                                            if header_column_index["product"] != -1:
                                                try:
                                                    consolidated_sheet_item.product_description = str(
                                                        columns[
                                                            header_column_index[
                                                                "product"
                                                            ]
                                                        ]
                                                    ).strip()
                                                except:
                                                    # May indicate a noisy row
                                                    pass
                                            if header_column_index["price"] != -1:
                                                try:
                                                    consolidated_sheet_item.unit_price = str(
                                                        columns[
                                                            header_column_index["price"]
                                                        ]
                                                    ).strip()
                                                    try:
                                                        consolidated_sheet_item.unit_price = float(
                                                            consolidated_sheet_item.unit_price
                                                        )
                                                    except:
                                                        pass
                                                except:
                                                    # May indicate a noisy row
                                                    pass

                                            # If lesser than 50%, it's noisy
                                            valid_threshold = 2
                                            valid_counter = 0
                                            if str(
                                                consolidated_sheet_item.barcode
                                            ).strip():
                                                valid_counter += 1
                                                if (
                                                    str(consolidated_sheet_item.barcode)
                                                    .strip()
                                                    .startswith("Unnamed:")
                                                ):
                                                    consolidated_sheet_item.barcode = ""
                                            if str(
                                                consolidated_sheet_item.quantity
                                            ).strip():
                                                valid_counter += 1
                                                if (
                                                    str(
                                                        consolidated_sheet_item.quantity
                                                    )
                                                    .strip()
                                                    .startswith("Unnamed:")
                                                ):
                                                    consolidated_sheet_item.quantity = (
                                                        ""
                                                    )
                                            if str(
                                                consolidated_sheet_item.product_description
                                            ).strip():
                                                valid_counter += 1
                                                if (
                                                    str(
                                                        consolidated_sheet_item.product_description
                                                    )
                                                    .strip()
                                                    .startswith("Unnamed:")
                                                ):
                                                    consolidated_sheet_item.product_description = (
                                                        ""
                                                    )
                                            if str(
                                                consolidated_sheet_item.unit_price
                                            ).strip():
                                                valid_counter += 1
                                                if (
                                                    str(
                                                        consolidated_sheet_item.unit_price
                                                    )
                                                    .strip()
                                                    .startswith("Unnamed:")
                                                ):
                                                    consolidated_sheet_item.unit_price = (
                                                        ""
                                                    )

                                            if valid_counter >= valid_threshold:
                                                consolidated_sheet_items.append(
                                                    consolidated_sheet_item
                                                )

                                            # For separate sheet
                                            try:
                                                required_columns = columns[
                                                    table_matrix_indices[
                                                        "left"
                                                    ] : table_matrix_indices["right"]
                                                    + 1
                                                ]

                                                # Checking index of column with real value despite knowing matrix to avoid noisy row
                                                lindex = 0
                                                rindex = 0
                                                for column_index in range(
                                                    len(required_columns)
                                                ):
                                                    if (
                                                        required_columns[column_index]
                                                        is not None
                                                        and str(
                                                            required_columns[
                                                                column_index
                                                            ]
                                                        ).strip()
                                                    ):
                                                        lindex = column_index
                                                        break
                                                for column_index in range(
                                                    len(required_columns) - 1, -1, -1
                                                ):
                                                    if (
                                                        required_columns[column_index]
                                                        is not None
                                                        and str(
                                                            required_columns[
                                                                column_index
                                                            ]
                                                        ).strip()
                                                    ):
                                                        rindex = column_index
                                                        break

                                                # If lesser than 50%, it's noisy
                                                if ((rindex + 1) - lindex) >= (
                                                    (table_matrix_indices["right"] + 1)
                                                    - table_matrix_indices["left"]
                                                ) / 2:
                                                    result_columns = []
                                                    for (
                                                        required_column
                                                    ) in required_columns:
                                                        if (
                                                            str(required_column)
                                                            .strip()
                                                            .startswith("Unnamed:")
                                                        ):
                                                            result_columns.append("")
                                                        else:
                                                            for column_index in range(
                                                                len(columns)
                                                            ):
                                                                if (
                                                                    str(
                                                                        columns[
                                                                            column_index
                                                                        ]
                                                                    ).strip()
                                                                    == str(
                                                                        required_column
                                                                    ).strip()
                                                                ):
                                                                    if (
                                                                        column_index
                                                                        == header_column_index[
                                                                            "barcode"
                                                                        ]
                                                                        or column_index
                                                                        == header_column_index[
                                                                            "quantity"
                                                                        ]
                                                                    ):
                                                                        try:
                                                                            result_columns.append(
                                                                                int(
                                                                                    float(
                                                                                        str(
                                                                                            required_column
                                                                                        ).strip()
                                                                                    )
                                                                                )
                                                                            )
                                                                        except:
                                                                            result_columns.append(
                                                                                required_column
                                                                            )
                                                                    elif (
                                                                        column_index
                                                                        == header_column_index[
                                                                            "quantity"
                                                                        ]
                                                                    ):
                                                                        try:
                                                                            result_columns.append(
                                                                                float(
                                                                                    str(
                                                                                        required_column
                                                                                    ).strip()
                                                                                )
                                                                            )
                                                                        except:
                                                                            result_columns.append(
                                                                                required_column
                                                                            )
                                                                    else:
                                                                        result_columns.append(
                                                                            required_column
                                                                        )
                                                                    break
                                                    separate_sheet.append(
                                                        result_columns
                                                    )
                                            except:
                                                # May indicate a noisy row
                                                pass

                                # Check if separate sheet has rows (including the header row)
                                if len(separate_sheet) > 1:
                                    separate_sheet_metadata.sheet = separate_sheet
                                    separate_sheets.append(separate_sheet_metadata)
                                # If not, remove the sheet name so that it doesn't appear in summary
                                else:
                                    separate_sheet_metadata.name = ""

                                # Update summary based on GPT analysis result
                                existence_count = len(
                                    [
                                        index
                                        for index in header_column_index.values()
                                        if index != -1
                                    ]
                                )
                                if existence_count == 4:
                                    report_sheet_item.status = "PROCESSED"
                                elif existence_count >= 2:
                                    report_sheet_item.status = "PARTIALLY PROCESSED"
                                    report_sheet_item.comments = (
                                        "Couldn't detect any headers that relate to"
                                    )
                                    if header_column_index["barcode"] == -1:
                                        report_sheet_item.comments += " barcode,"
                                    if header_column_index["quantity"] == -1:
                                        report_sheet_item.comments += " quantity,"
                                    if header_column_index["product"] == -1:
                                        report_sheet_item.comments += (
                                            " product description,"
                                        )
                                    if header_column_index["price"] == -1:
                                        report_sheet_item.comments += " unit price,"
                                    report_sheet_item.comments = (
                                        report_sheet_item.comments[
                                            0 : len(report_sheet_item.comments) - 1
                                        ]
                                    )
                                else:
                                    report_sheet_item.status = "NOT PROCESSED"
                                    if existence_count == 1:
                                        report_sheet_item.comments = "Could detect just one header that relate to barcode, quantity, product description or unit price."
                                    else:
                                        report_sheet_item.comments = "Couldn't detect any headers that relate to barcode, quantity, product description, unit price."
                                    report_sheet_item[2] = ""
                                report_sheet_items.append(report_sheet_item)
                except Exception as ex:
                    system_logger.info(
                        "\nException occurred while processing the mail from:\nSender: "
                        + message.sender.name
                        + " - "
                        + message.sender.address
                        + "\n"
                        + "Received at: "
                        + str(message.received_at)
                        + "\n"
                        + "Attachment failed to process: "
                        + attachment.name
                        + "\n"
                    )
                    system_logger.exception(ex)
                    report_sheet_items.append(
                        ReportSheetItem(
                            sender=message.sender.name + " - " + message.sender.address,
                            received_at=message.received_at,
                            sheet_name="",
                            status="NOT PROCESSED",
                            comments="System Exception: Couldn't process the attachment. Check logs for more details.",
                            file_name=attachment.name,
                            file_path=f"{os.path.join(output_directory, 'Messages', 'Message ' + str(message_counter), 'Attachments', attachment.name)}",
                        )
                    )

    # PriceRunner
    get_price_results(consolidated_sheet_items, config, system_logger)

    # Generate report
    generate_report(
        output_directory, separate_sheets, consolidated_sheet_items, report_sheet_items
    )


def initiate_assistant():
    assistant = client.beta.assistants.create(
        name="Sales Data Analyser",
        instructions='You are a sales data assistant who identifies the header row (i.e., The CSV line) and provides four headers from a CSV (Comma Separated Values) file content in the JSON dict format: {"barcode": <Barcode_Column>, "quantity": <Quantity_Column>, "product": <Product_Column>, "price": <Price_Column>}. Each property relates to Barcode, Quantity, Product description and Price column headers respectively in the CSV file. The synonym of the header\'s words, or sometimes abbreviations, should match the requirement. For example, the header for Barcode (a GTIN) could be EAN, UPC, GTIN, etc. The header for Quantity could be Quantity, Stock, Pieces, PCS, QTY, Units, etc. The header for Price could be Price, Unit Price, Cost, etc. The header for Product could be Product, Name, Description, Item, DESC, etc. The result values should strictly be from the same row which is considered as header row and not different rows. You could also look the the column\'s values to verify if most of them relate to what you decide the column\'s header is. For example, you can verify the column header for Barcode by not only looking at the header\'s words, but also if most of that column\'s values has valid barcodes (GTIN). If you could not find a suitable value for a property, set an empty string to that.',
        # model="ft:gpt-3.5-turbo-0125:vallabha-systems-limited:header-prediction:9lFHETSg"
        # Latest
        model="ft:gpt-3.5-turbo-0125:vallabha-systems-limited:header-prediction:9nk6uHeF",
    )
    return assistant


def initialize_thread():
    thread = client.beta.threads.create()
    return thread


def reinitialize_thread(thread: Thread):
    for message in client.beta.threads.messages.list(thread_id=thread.id):
        try:
            client.beta.threads.messages.delete(
                thread_id=thread.id, message_id=message.id
            )
        except:
            pass
        return client.beta.threads.retrieve(thread_id=thread.id)


def analyse(prompt_content: str, thread: Thread, assistant: Assistant):
    if thread is None:
        thread = initialize_thread()
    else:
        thread = reinitialize_thread(thread)

    client.beta.threads.messages.create(
        thread_id=thread.id,
        role="user",
        content="Please provide me the four headers from the header row that relate to Barcode, Product, Quantity and Price in a JSON dict format by analysing the given CSV content:\n"
        + prompt_content,
    )

    run = client.beta.threads.runs.create_and_poll(
        assistant_id=assistant.id, thread_id=thread.id
    )
    return thread, run


def calculate_table_matrix_indices(table_matrix_indices, csv_rows, header_column_index):
    left_indices = []
    right_indices = []
    for row_index in range(table_matrix_indices["top"] + 1, len(csv_rows)):
        columns = [str(col).strip() for col_idx, col in enumerate(csv_rows[row_index])]

        for column_index in range(len(columns)):
            if column_index in header_column_index.values() or (
                columns[column_index] is not None and str(columns[column_index]).strip()
            ):
                left_indices.append(column_index)
                break
        for column_index in range(len(columns) - 1, -1, -1):
            if column_index in header_column_index.values() or (
                columns[column_index] is not None and str(columns[column_index]).strip()
            ):
                right_indices.append(column_index)
                break

    table_matrix_indices["left"] = statistics.mode(left_indices)
    table_matrix_indices["right"] = statistics.mode(right_indices)
