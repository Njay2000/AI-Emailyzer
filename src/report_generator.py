import os
from typing import List

from openpyxl.styles import PatternFill, Font
from openpyxl.workbook import Workbook

from src.models.consolidated_sheet import ConsolidatedSheetItem
from src.models.report_sheet_item import ReportSheetItem
from src.models.separate_sheet_metadata import SeparateSheetMetadata


def generate_report(
    output_directory: str,
    separate_sheets: List[SeparateSheetMetadata],
    consolidated_sheet_items: List[ConsolidatedSheetItem],
    report_sheet_items: List[ReportSheetItem],
):
    print("⏳ Generating report")
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    ws = wb.create_sheet(title="Summary")
    ws.append(["SENDER", "TIME", "SHEET", "STATUS", "COMMENTS", "FILE"])
    for row_index in range(len(report_sheet_items)):
        ws.append(
            [
                report_sheet_items[row_index].sender,
                report_sheet_items[row_index].received_at,
                report_sheet_items[row_index].sheet_name,
                report_sheet_items[row_index].status,
                report_sheet_items[row_index].comments,
                report_sheet_items[row_index].file_name,
            ]
        )

        if str(report_sheet_items[row_index].sheet_name).strip():
            add_hyperlink(
                worksheet=ws,
                row=row_index + 2,
                column=3,
                link=f"#'{str(report_sheet_items[row_index].sheet_name).strip()}'!A1",
            )
        add_hyperlink(
            worksheet=ws,
            row=row_index + 2,
            column=6,
            link=report_sheet_items[row_index].file_path,
        )
    adjust_column_style(ws)
    update_status(ws)

    ws = wb.create_sheet(title="Consolidated")
    ws.append(
        [
            "SENDER",
            "BARCODE",
            "QUANTITY",
            "PRODUCT DESCRIPTION",
            "UNIT PRICE",
            "PRICERUNNER - MEDIAN PRICE",
            "PRICERUNNER - LOWEST PRICE",
            "PRICERUNNER - RETAILER (LOWEST PRICE)",
            "PRICERUNNER - HIGHEST PRICE",
            "PRICERUNNER - RETAILER (HIGHEST PRICE)",
            "PRICERUNNER - AVERAGE PRICE",
        ]
    )
    for row in consolidated_sheet_items:
        ws.append(
            [
                row.sender,
                row.barcode,
                row.quantity,
                row.product_description,
                row.unit_price,
                row.price_runner_details.median,
                row.price_runner_details.lowest_price,
                row.price_runner_details.lowest_price_retailer,
                row.price_runner_details.highest_price,
                row.price_runner_details.highest_price_retailer,
                row.price_runner_details.average_price,
            ]
        )
    adjust_column_style(ws)

    for sheet in separate_sheets:
        ws = wb.create_sheet(title=sheet.name)

        for row in sheet.sheet:
            ws.append(row)
        adjust_column_style(ws)

    wb.save(os.path.join(output_directory, "Report.xlsx"))


def adjust_column_style(worksheet):
    for col in worksheet.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 1
        column_letter = col[0].column_letter
        worksheet.column_dimensions[column_letter].width = adjusted_width
        worksheet[column_letter + "1"].font = Font(bold=True)


def update_status(worksheet):
    yellow_fill = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    green_fill = PatternFill(
        start_color="90EE90", end_color="90EE90", fill_type="solid"
    )
    red_fill = PatternFill(
        start_color="FF6347", end_color="FF6347", fill_type="solid"
    )  # Red color

    for row in worksheet.iter_rows(
        min_col=4, max_col=4, min_row=2, max_row=worksheet.max_row
    ):
        for cell in row:
            if str(cell.value).strip() == "PROCESSED":
                cell.fill = green_fill
            elif str(cell.value).strip() == "PARTIALLY PROCESSED":
                cell.fill = yellow_fill
            else:
                cell.fill = red_fill


def add_hyperlink(worksheet, row, column, link):
    cell = worksheet.cell(row=row, column=column)
    cell.hyperlink = link
    cell.font = Font(color="0000FF", underline="single")
